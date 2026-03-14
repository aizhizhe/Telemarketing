from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from .models import KnowledgeRecord, PainPoint, RetrievalHit


QUESTION_MARKERS = (
    "?",
    "？",
    "吗",
    "怎么",
    "为什么",
    "如何",
    "多少",
    "哪里",
    "哪儿",
    "哪科",
    "价格",
    "报价",
    "费用",
    "校区",
    "线上",
    "试课",
    "试听",
    "微信",
)

OBJECTION_MARKERS = (
    "没时间",
    "太贵",
    "太远",
    "不需要",
    "不想",
    "没用",
    "先不用",
    "不考虑",
)

FLOW_ONLY_MARKERS = (
    "可以",
    "行",
    "好",
    "好的",
    "嗯",
    "加微信",
    "留微信",
    "留电话",
    "安排",
)

WEAK_OBJECTION_KEYWORDS = {"担心"}
SUBJECT_ALIASES = {
    "语文": ("语文",),
    "数学": ("数学",),
    "英语": ("英语",),
    "物理": ("物理",),
    "化学": ("化学",),
    "生物": ("生物",),
    "历史": ("历史",),
    "地理": ("地理",),
    "政治": ("政治", "道法", "道德与法治"),
}
GRADE_ALIAS_MAP = {
    "一年级": "小学1年级",
    "二年级": "小学2年级",
    "三年级": "小学3年级",
    "四年级": "小学4年级",
    "五年级": "小学5年级",
    "六年级": "小学6年级",
    "初一": "七年级",
    "初二": "八年级",
    "初三": "九年级",
}
PAIN_GRADE_CANDIDATES = {
    "小学1年级": ["小学1年级", "小学1-4年级"],
    "小学2年级": ["小学2年级", "小学1-4年级"],
    "小学3年级": ["小学3年级", "小学1-4年级"],
    "小学4年级": ["小学4年级", "小学1-4年级"],
    "小学5年级": ["小学5年级"],
    "小学6年级": ["小学6年级"],
    "七年级": ["七年级", "初一"],
    "八年级": ["八年级", "初二"],
    "九年级": ["九年级", "初三"],
    "高一": ["高一"],
    "高二": ["高二"],
    "高三": ["高三"],
}
GRADE_PATTERNS = (
    (re.compile(r"(小学[1-6一二三四五六]年级)"), None),
    (re.compile(r"([一二三四五六]年级)"), None),
    (re.compile(r"(初[一二三])"), None),
    (re.compile(r"([七八九]年级)"), None),
    (re.compile(r"(高[一二三])"), None),
    (re.compile(r"([1-6])年级"), "小学{n}年级"),
)
PHONE_REGEX = re.compile(r"(?<!\d)(1[3-9]\d{9})(?!\d)")
WECHAT_REGEX = re.compile(r"(?<![A-Za-z0-9_])([A-Za-z][A-Za-z0-9_-]{5,19})(?![A-Za-z0-9_])")
ORDER_REGEX = re.compile(r"(?:订单号|单号|工单号)[:： ]*([A-Za-z0-9-]{6,})")
NAME_REGEXES = (
    re.compile(r"(?:我叫|我是|姓)([\u4e00-\u9fff]{1,4})"),
    re.compile(r"([\u4e00-\u9fff]{2,4})(?:老师|女士|先生|家长)"),
)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _normalize_lookup_text(text: str) -> str:
    return re.sub(r"[^\w\u4e00-\u9fff]+", "", text).lower()


def _build_text_ngrams(text: str, min_n: int = 2, max_n: int = 3) -> set[str]:
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "", str(text or ""))
    grams: set[str] = set()
    for n in range(min_n, max_n + 1):
        if len(cleaned) < n:
            continue
        for index in range(len(cleaned) - n + 1):
            grams.add(cleaned[index : index + n])
    return grams


def _split_question_variants(question: str) -> list[str]:
    return [part.strip() for part in re.split(r"[/／]", str(question or "")) if part.strip()]


def _split_keywords(value: object) -> tuple[str, ...]:
    text = _clean_text(value)
    if not text:
        return ()
    parts = re.split(r"[、，,/\s]+", text)
    return tuple(part.strip() for part in parts if part and part.strip())


def _canonical_grade(value: str) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    text = GRADE_ALIAS_MAP.get(text, text)
    return (
        text.replace("小学一年级", "小学1年级")
        .replace("小学二年级", "小学2年级")
        .replace("小学三年级", "小学3年级")
        .replace("小学四年级", "小学4年级")
        .replace("小学五年级", "小学5年级")
        .replace("小学六年级", "小学6年级")
    )


def extract_grade(text: str) -> str | None:
    cleaned = _clean_text(text)
    if not cleaned:
        return None
    for pattern, template in GRADE_PATTERNS:
        match = pattern.search(cleaned)
        if not match:
            continue
        if template:
            return template.format(n=match.group(1))
        return _canonical_grade(match.group(1))
    return None


def extract_subjects(text: str) -> list[str]:
    cleaned = _clean_text(text)
    subjects: list[str] = []
    for canonical, aliases in SUBJECT_ALIASES.items():
        if any(alias in cleaned for alias in aliases):
            subjects.append(canonical)
    return subjects


def extract_phone(text: str) -> str | None:
    match = PHONE_REGEX.search(_clean_text(text))
    return match.group(1) if match else None


def extract_wechat(text: str) -> str | None:
    cleaned = _clean_text(text)
    if "微信" not in cleaned and "vx" not in cleaned.lower() and "wx" not in cleaned.lower():
        return None
    match = WECHAT_REGEX.search(cleaned)
    if not match:
        return None
    candidate = match.group(1)
    if candidate.lower() in {"wechat", "weixin", "hello", "thanks"}:
        return None
    return candidate


def extract_order_no(text: str) -> str | None:
    match = ORDER_REGEX.search(_clean_text(text))
    return match.group(1) if match else None


def extract_name(text: str) -> str | None:
    cleaned = _clean_text(text)
    for pattern in NAME_REGEXES:
        match = pattern.search(cleaned)
        if match:
            return match.group(1)
    return None


class KnowledgeBase:
    def __init__(self, base_dir: Path) -> None:
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.qa_records = self._load_qa_records()
        self.objection_records = self._load_objection_records()
        self.pain_points = self._load_pain_points()
        self.record_count = len(self.qa_records) + len(self.objection_records)

    def _workbooks(self) -> list[Path]:
        return sorted(self.base_dir.glob("*.xlsx"))

    def _find_all_workbooks(self, *keywords: str) -> list[Path]:
        matches: list[Path] = []
        for path in self._workbooks():
            if all(keyword in path.name for keyword in keywords):
                matches.append(path)
        return matches

    def _find_first_workbook(self, *keywords: str) -> Path:
        matches = self._find_all_workbooks(*keywords)
        if not matches:
            raise FileNotFoundError(f"Workbook not found for keywords={keywords}")
        return matches[0]

    @staticmethod
    def _sheet_rows(path: Path) -> list[list[str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows: list[list[str]] = []
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for raw_row in worksheet.iter_rows(values_only=True):
                    rows.append([_clean_text(cell) for cell in raw_row])
        finally:
            workbook.close()
        return rows

    @staticmethod
    def _find_header_row(rows: Sequence[Sequence[str]], required_tokens: Sequence[str]) -> tuple[int, list[str]]:
        for index, row in enumerate(rows[:10]):
            if all(any(token in cell for cell in row) for token in required_tokens):
                return index, list(row)
        raise ValueError(f"Header row not found for tokens={required_tokens}")

    @staticmethod
    def _rows_to_dicts(rows: Sequence[Sequence[str]], required_tokens: Sequence[str]) -> list[dict[str, str]]:
        header_index, header = KnowledgeBase._find_header_row(rows, required_tokens)
        mapped_rows: list[dict[str, str]] = []
        for row in rows[header_index + 1 :]:
            if not any(cell for cell in row):
                continue
            values = list(row) + [""] * max(0, len(header) - len(row))
            mapped_rows.append({header[i]: values[i] for i in range(len(header))})
        return mapped_rows

    @staticmethod
    def _pick_value(row: dict[str, str], *tokens: str) -> str:
        for key, value in row.items():
            if any(token in key for token in tokens):
                return _clean_text(value)
        return ""

    @staticmethod
    def _dedupe(records: Iterable[KnowledgeRecord], key_fn) -> list[KnowledgeRecord]:
        seen: set[tuple[object, ...]] = set()
        deduped: list[KnowledgeRecord] = []
        for record in records:
            key = key_fn(record)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(record)
        return deduped

    def _load_qa_records(self) -> list[KnowledgeRecord]:
        path = self._find_first_workbook("问答知识库")
        rows = self._rows_to_dicts(
            self._sheet_rows(path),
            required_tokens=("问题", "适用场景", "答案"),
        )
        records: list[KnowledgeRecord] = []
        for row in rows:
            question = self._pick_value(row, "问题")
            scene = self._pick_value(row, "适用场景")
            answer = self._pick_value(row, "答案")
            if not question or not answer:
                continue
            text = "\n".join(
                [
                    "库类型: 问答知识库",
                    f"问题: {question}",
                    f"适用场景: {scene}",
                    f"答案: {answer}",
                ]
            )
            records.append(
                KnowledgeRecord(
                    kb_type="qa",
                    title=question,
                    question=question,
                    answer=answer,
                    scene=scene,
                    keywords=(),
                    source_file=path.name,
                    text=text,
                )
            )
        return self._dedupe(records, lambda item: (item.kb_type, item.title, item.answer, item.scene))

    def _load_objection_records(self) -> list[KnowledgeRecord]:
        results: list[KnowledgeRecord] = []
        for path in self._find_all_workbooks("异议答题库"):
            rows = self._rows_to_dicts(
                self._sheet_rows(path),
                required_tokens=("素材标签", "触发关键词", "素材内容"),
            )
            for row in rows:
                title = self._pick_value(row, "素材标签")
                scene = self._pick_value(row, "触发场景")
                answer = self._pick_value(row, "素材内容")
                keywords = _split_keywords(self._pick_value(row, "触发关键词"))
                if not title or not answer:
                    continue
                text = "\n".join(
                    [
                        "库类型: 异议答题库",
                        f"标签: {title}",
                        f"场景: {scene}",
                        f"关键词: {'、'.join(keywords)}",
                        f"话术: {answer}",
                    ]
                )
                results.append(
                    KnowledgeRecord(
                        kb_type="objection",
                        title=title,
                        question=scene or title,
                        answer=answer,
                        scene=scene,
                        keywords=keywords,
                        source_file=path.name,
                        text=text,
                    )
                )
        return self._dedupe(results, lambda item: (item.kb_type, item.title, item.answer, item.scene))

    def _load_pain_points(self) -> list[PainPoint]:
        path = self._find_first_workbook("痛点")
        rows = self._rows_to_dicts(
            self._sheet_rows(path),
            required_tokens=("年级", "学科", "学科痛点"),
        )
        current_grade = ""
        points: list[PainPoint] = []
        for row in rows:
            grade = _canonical_grade(self._pick_value(row, "年级"))
            subject = self._pick_value(row, "学科")
            content = self._pick_value(row, "学科痛点")
            if grade:
                current_grade = grade
            if not subject or not content:
                continue
            points.append(PainPoint(grade=current_grade, subject=subject, content=content))
        return points

    def should_use_rag(self, text: str) -> bool:
        cleaned = _clean_text(text)
        if not cleaned:
            return False
        if any(marker in cleaned for marker in QUESTION_MARKERS):
            return True
        if any(marker in cleaned for marker in OBJECTION_MARKERS):
            return True
        if len(cleaned) <= 8 and any(marker == cleaned for marker in FLOW_ONLY_MARKERS):
            return False
        return any(subject in cleaned for subject in SUBJECT_ALIASES) or "校区" in cleaned or "试听" in cleaned

    def search(self, query: str, top_k: int = 8, preferred_type: str | None = None) -> list[RetrievalHit]:
        hits: list[RetrievalHit] = []
        if preferred_type in (None, "qa"):
            hits.extend(self.search_qa(query))
        if preferred_type in (None, "objection"):
            hits.extend(self.search_objection(query))
        deduped: dict[tuple[str, str, str], RetrievalHit] = {}
        for hit in hits:
            key = (hit.kb_type, hit.title, hit.answer)
            current = deduped.get(key)
            if current is None or hit.score > current.score:
                deduped[key] = hit
        return sorted(deduped.values(), key=lambda item: item.score, reverse=True)[:top_k]

    def search_qa(self, user_text: str) -> list[RetrievalHit]:
        user_text = _clean_text(user_text)
        if not user_text:
            return []
        user_norm = _normalize_lookup_text(user_text)
        user_grams = _build_text_ngrams(user_text)
        matches: list[RetrievalHit] = []
        for qa in self.qa_records:
            variants = _split_question_variants(qa.question)
            best_score = 0.0
            best_variant = ""
            for variant in variants:
                variant_norm = _normalize_lookup_text(variant)
                if not variant_norm:
                    continue
                score = 0.0
                if user_norm == variant_norm:
                    score += 12.0
                elif variant_norm in user_norm:
                    score += 9.0
                elif len(user_norm) >= 4 and user_norm in variant_norm:
                    score += 7.0
                overlap_count = len(user_grams & _build_text_ngrams(variant))
                score += min(overlap_count, 10) * 0.5
                token_hits = 0
                for token in re.split(r"[、，, ]+", variant.replace("？", "").replace("?", "")):
                    if len(token) >= 2 and token in user_text:
                        token_hits += 1
                score += min(token_hits, 4) * 0.8
                if score > best_score:
                    best_score = score
                    best_variant = variant
            if best_score < 2.5:
                continue
            matches.append(
                RetrievalHit(
                    kb_type=qa.kb_type,
                    title=qa.title,
                    question=qa.question,
                    answer=qa.answer,
                    scene=qa.scene,
                    keywords=qa.keywords,
                    source_file=qa.source_file,
                    text=qa.text,
                    score=best_score,
                    matched_variant=best_variant,
                )
            )
        return sorted(matches, key=lambda item: (item.score, len(item.answer)), reverse=True)

    def search_objection(self, user_text: str) -> list[RetrievalHit]:
        user_text = _clean_text(user_text)
        if not user_text:
            return []
        matches: list[RetrievalHit] = []
        user_grams = _build_text_ngrams(user_text)
        for record in self.objection_records:
            matched_keywords = [kw for kw in record.keywords if kw and kw in user_text]
            if not matched_keywords:
                continue
            if len(matched_keywords) == 1 and matched_keywords[0] in WEAK_OBJECTION_KEYWORDS:
                continue
            score = 0.0
            for keyword in matched_keywords:
                if keyword in WEAK_OBJECTION_KEYWORDS:
                    score += 0.2
                else:
                    score += 1.0 + min(len(keyword), 4) * 0.1
            overlap_count = len(user_grams & _build_text_ngrams(record.text))
            score += min(overlap_count, 8) * 0.15
            matches.append(
                RetrievalHit(
                    kb_type=record.kb_type,
                    title=record.title,
                    question=record.question,
                    answer=record.answer,
                    scene=record.scene,
                    keywords=record.keywords,
                    source_file=record.source_file,
                    text=record.text,
                    score=score,
                    matched_variant="、".join(matched_keywords),
                )
            )
        return sorted(matches, key=lambda item: item.score, reverse=True)

    def get_pain_point(self, grade: str | None, subject: str | None) -> PainPoint | None:
        canonical_grade = _canonical_grade(grade or "")
        canonical_subject = _clean_text(subject)
        if not canonical_grade or not canonical_subject:
            return None
        candidates = PAIN_GRADE_CANDIDATES.get(canonical_grade, [canonical_grade])
        for candidate in candidates:
            for point in self.pain_points:
                if point.grade == candidate and point.subject == canonical_subject:
                    return point
        return None

    def knowledge_documents(self) -> list[dict[str, str]]:
        documents: list[dict[str, str]] = []
        for record in [*self.qa_records, *self.objection_records]:
            documents.append(
                {
                    "doc_title": record.title,
                    "doc_type": record.kb_type,
                    "content_text": record.text,
                    "source_file": record.source_file,
                }
            )
        return documents
