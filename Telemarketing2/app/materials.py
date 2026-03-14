from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import re

from openpyxl import load_workbook

from .config import PROJECT_ROOT


DATA_DIR = PROJECT_ROOT / "Data"
DATA2_DIR = PROJECT_ROOT / "Data2"

SUBJECT_ALIASES = {
    "语文": ("语文",),
    "数学": ("数学",),
    "英语": ("英语",),
    "物理": ("物理",),
    "化学": ("化学",),
    "生物": ("生物",),
    "历史": ("历史",),
    "地理": ("地理",),
    "政治": ("政治", "道法", "道德与法治"),
}

GRADE_DIRECT_MAP = {
    "一年级": "小学1年级",
    "二年级": "小学2年级",
    "三年级": "小学3年级",
    "四年级": "小学4年级",
    "五年级": "小学5年级",
    "六年级": "小学6年级",
    "初一": "七年级",
    "初二": "八年级",
    "初三": "九年级",
}

GRADE_PATTERNS: list[tuple[re.Pattern[str], str | None]] = [
    (re.compile(r"(小学[1-6]年级)"), None),
    (re.compile(r"([一二三四五六]年级)"), None),
    (re.compile(r"(初一|初二|初三|高一|高二|高三)"), None),
    (re.compile(r"(七年级|八年级|九年级)"), None),
]

GRADE_CANDIDATES = {
    "小学1年级": ("小学1年级", "一年级", "小学1-4年级"),
    "小学2年级": ("小学2年级", "二年级", "小学1-4年级"),
    "小学3年级": ("小学3年级", "三年级", "小学1-4年级"),
    "小学4年级": ("小学4年级", "四年级", "小学1-4年级"),
    "小学5年级": ("小学5年级", "五年级"),
    "小学6年级": ("小学6年级", "六年级"),
    "七年级": ("七年级", "初一"),
    "八年级": ("八年级", "初二"),
    "九年级": ("九年级", "初三"),
    "高一": ("高一",),
    "高二": ("高二",),
    "高三": ("高三",),
}

WECHAT_REGEX = re.compile(r"(?<![A-Za-z0-9_])(1[3-9]\d{9}|[A-Za-z][A-Za-z0-9_-]{5,19})(?![A-Za-z0-9_])")


@dataclass(frozen=True)
class ProfessionalMaterial:
    material_id: str
    source_file: str
    category: str
    title: str
    scene: str
    content: str
    keywords: tuple[str, ...]
    grade: str | None = None
    subject: str | None = None


@dataclass(frozen=True)
class FlowMaterial:
    material_id: str
    node: str
    tag: str
    scene: str
    content: str


@dataclass(frozen=True)
class FlowRule:
    rule_id: str
    target: str
    logic: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").replace("\n", " ").strip()


def split_keywords(value: str) -> tuple[str, ...]:
    text = clean_text(value)
    if not text:
        return ()
    parts = re.split(r"[、，,/\s]+", text)
    return tuple(part for part in (item.strip() for item in parts) if part)


def extract_grade(text: str) -> str | None:
    cleaned = clean_text(text)
    if not cleaned:
        return None
    for pattern, template in GRADE_PATTERNS:
        match = pattern.search(cleaned)
        if not match:
            continue
        value = match.group(1)
        value = GRADE_DIRECT_MAP.get(value, value)
        return template.format(n=value) if template else value
    return None


def extract_subjects(text: str) -> list[str]:
    cleaned = clean_text(text)
    results: list[str] = []
    for subject, aliases in SUBJECT_ALIASES.items():
        if any(alias in cleaned for alias in aliases):
            results.append(subject)
    return results


def extract_wechat(text: str) -> str | None:
    cleaned = clean_text(text)
    if not cleaned:
        return None
    if "手机号同微信" in cleaned or ("手机" in cleaned and "微信" in cleaned and "就是" in cleaned):
        return "手机号同微信"
    match = WECHAT_REGEX.search(cleaned)
    if not match:
        return None
    candidate = match.group(1)
    if candidate.lower() in {"wechat", "weixin", "wx", "hello"}:
        return None
    return candidate


def split_sentences(text: str, limit: int = 2) -> str:
    content = clean_text(text)
    if not content:
        return ""
    parts = re.split(r"(?<=[。！？!?])", content)
    sentences = [item.strip() for item in parts if item.strip()]
    if not sentences:
        return content
    return "".join(sentences[:limit]).strip()


def grade_candidates(grade: str | None) -> tuple[str, ...]:
    if not grade:
        return ()
    return GRADE_CANDIDATES.get(grade, (grade,))


class MaterialRepository:
    def __init__(self, project_root: Path | None = None) -> None:
        self.project_root = project_root or PROJECT_ROOT
        self.professional_materials = self._load_professional_materials()
        self.flow_materials = self._load_flow_materials()
        self.flow_rules = self._load_flow_rules()

    def _iter_rows(self, workbook_path: Path, sheet_name: str = "Sheet1") -> list[dict[str, str]]:
        wb = load_workbook(workbook_path, read_only=True)
        ws = wb[sheet_name]
        raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [clean_text(value) for value in raw_headers]
        rows: list[dict[str, str]] = []
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[idx]: clean_text(value) for idx, value in enumerate(raw_row) if idx < len(headers)}
            if any(item.values()):
                rows.append(item)
        return rows

    def _load_professional_materials(self) -> list[ProfessionalMaterial]:
        materials: list[ProfessionalMaterial] = []

        qa_rows = self._iter_rows(DATA_DIR / "问答知识库1.0版.xlsx")
        for idx, row in enumerate(qa_rows, start=1):
            materials.append(
                ProfessionalMaterial(
                    material_id=f"qa-{idx}",
                    source_file="问答知识库1.0版.xlsx",
                    category="qa",
                    title=row.get("问题", ""),
                    scene=row.get("适用场景", ""),
                    content=row.get("答案", ""),
                    keywords=split_keywords(row.get("问题", "")),
                )
            )

        objection_rows = self._iter_rows(DATA_DIR / "异议答题库(1).xlsx")
        for idx, row in enumerate(objection_rows, start=1):
            materials.append(
                ProfessionalMaterial(
                    material_id=f"objection-{idx}",
                    source_file="异议答题库(1).xlsx",
                    category="objection",
                    title=row.get("素材标签（含核心原则）", ""),
                    scene=row.get("触发场景", ""),
                    content=row.get("素材内容（带拉回主线）", ""),
                    keywords=split_keywords(row.get("触发关键词", "")),
                )
            )

        pain_rows = self._iter_rows(DATA_DIR / "素材库-年级-学科痛点1.0版本.xlsx")
        for idx, row in enumerate(pain_rows, start=1):
            grade = row.get("年级", "") or None
            subject = row.get("学科", "") or None
            content = row.get("学科痛点", "")
            if not content:
                continue
            materials.append(
                ProfessionalMaterial(
                    material_id=f"pain-{idx}",
                    source_file="素材库-年级-学科痛点1.0版本.xlsx",
                    category="pain_point",
                    title=f"{grade or '通用'}-{subject or '通用'}",
                    scene="年级学科痛点",
                    content=content,
                    keywords=tuple(part for part in (grade, subject) if part),
                    grade=grade,
                    subject=subject,
                )
            )
        return materials

    def _load_flow_materials(self) -> list[FlowMaterial]:
        rows = self._iter_rows(DATA2_DIR / "节点素材库.xlsx")
        return [
            FlowMaterial(
                material_id=f"flow-{idx}",
                tag=row.get("素材标签", ""),
                node=row.get("适配节点", ""),
                scene=row.get("触发场景", ""),
                content=row.get("素材内容", ""),
            )
            for idx, row in enumerate(rows, start=1)
            if row.get("素材内容")
        ]

    def _load_flow_rules(self) -> list[FlowRule]:
        rows = self._iter_rows(DATA2_DIR / "RAG触发规则表.xlsx")
        rules: list[FlowRule] = []
        for idx, row in enumerate(rows, start=1):
            logic = row.get("触发逻辑", "")
            if not logic:
                continue
            rules.append(
                FlowRule(
                    rule_id=f"rule-{idx}",
                    target=row.get("适配节点/库", ""),
                    logic=logic,
                )
            )
        return rules

    def stats(self) -> dict[str, int]:
        return {
            "professional_materials": len(self.professional_materials),
            "flow_materials": len(self.flow_materials),
            "flow_rules": len(self.flow_rules),
        }

    def rules_for_target(self, target: str) -> list[FlowRule]:
        return [rule for rule in self.flow_rules if target in rule.target or rule.target in target]

    def search_flow(self, *, node: str, preferred_tags: list[str] | None = None, top_k: int = 4) -> list[dict]:
        tags = preferred_tags or []
        scored: list[tuple[float, FlowMaterial]] = []
        for material in self.flow_materials:
            if material.node != node:
                continue
            score = 1.0
            if any(tag and tag in material.tag for tag in tags):
                score += 4.0
            if any(tag and tag in material.content for tag in tags):
                score += 1.0
            scored.append((score, material))
        scored.sort(key=lambda item: item[0], reverse=True)
        return [
            {
                "material_id": material.material_id,
                "node": material.node,
                "tag": material.tag,
                "scene": material.scene,
                "content": material.content,
                "score": score,
            }
            for score, material in scored[:top_k]
        ]

    def search_professional(
        self,
        *,
        query: str,
        objection_label: str | None,
        professional_topics: list[str],
        grade: str | None,
        subject: str | None,
        top_k: int = 5,
    ) -> tuple[list[dict], bool]:
        cleaned_query = clean_text(query)
        query_terms: list[str] = [term for term in professional_topics if term]
        if objection_label:
            query_terms.append(objection_label)
        for term in ("成绩", "提分", "薄弱", "试听", "老师", "匹配", "顾问", "校区", "线上", "位置", "费用", "报价", "原因", "方式", "关键", "重点"):
            if term in cleaned_query:
                query_terms.append(term)
        query_terms = list(dict.fromkeys(query_terms))
        scored: list[tuple[float, ProfessionalMaterial]] = []
        for material in self.professional_materials:
            score = 0.0
            haystacks = [material.title, material.scene, material.content]
            strong_pain_match = False
            if material.category == "pain_point":
                strong_pain_match = bool(
                    grade
                    and subject
                    and material.subject == subject
                    and material.grade
                    and any(candidate in material.grade for candidate in grade_candidates(grade))
                )
                if not strong_pain_match:
                    continue
            if cleaned_query and any(cleaned_query in haystack for haystack in haystacks):
                score += 8.0
            for keyword in material.keywords:
                if keyword and keyword in cleaned_query:
                    score += 2.5
            for term in query_terms:
                if term and term in material.title:
                    score += 2.5
                elif term and term in material.scene:
                    score += 1.5
                elif term and term in material.content:
                    score += 1.0
            if objection_label and objection_label in "".join(haystacks):
                score += 2.5
            for topic in professional_topics:
                if topic and any(topic in haystack for haystack in haystacks):
                    score += 1.5
            if material.category == "pain_point" and strong_pain_match:
                score += 5.0
            if subject and material.subject == subject:
                score += 2.0
            if grade and material.grade and any(candidate in material.grade for candidate in grade_candidates(grade)):
                score += 1.5
            if material.category == "objection" and objection_label:
                score += 1.0
            if score > 0:
                scored.append((score, material))

        scored.sort(key=lambda item: item[0], reverse=True)
        hits = [
            {
                "material_id": material.material_id,
                "source_file": material.source_file,
                "category": material.category,
                "title": material.title,
                "scene": material.scene,
                "content": material.content,
                "score": score,
            }
            for score, material in scored[:top_k]
        ]
        gap_detected = not hits or hits[0]["score"] < 3.0
        return hits, gap_detected


@lru_cache(maxsize=1)
def get_material_repository() -> MaterialRepository:
    return MaterialRepository()
